"""Excel writer with auto-formatting for financial data."""

import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


class ExcelWriter:
    """
    Excel writer that automatically formats columns and prevents scientific notation.

    Features:
    - Auto-adjusts column widths based on header and data content
    - Formats monetary columns with thousands separators
    - Displays negative numbers in parentheses
    - Prevents scientific notation for large numbers
    - Appends to existing files, preserving other sheets
    - Replaces sheet if same name already exists
    - Header row styling (bold, blue background, white text)
    - Freeze panes at header row for easy scrolling
    - Auto-filter enabled on all data columns

    Attributes:
        path: File path for the Excel output
        writer: pandas ExcelWriter instance
    """

    # Column width constants
    MIN_COLUMN_WIDTH = 14  # Minimum column width in characters
    MAX_COLUMN_WIDTH = 75  # Maximum column width in characters
    COLUMN_PADDING = 2  # Extra padding added to column width

    # Number format strings for Excel
    MONETARY_FORMAT = "#,##0.00;(#,##0.00)"  # Currency with 2 decimals, comma separator, negatives in parentheses
    NUMERIC_FORMAT = "#,##0"  # Integer with comma separator

    # Keywords to identify monetary columns
    MONETARY_KEYWORDS = ["amount", "balance", "total", "credit", "debit"]

    def __init__(self, path: str) -> None:
        """
        Initialize ExcelWriter with output path.

        Args:
            path: File path where Excel file will be saved

        Note:
            Creates parent directories if they don't exist.
            Opens file in append mode if it exists, preserving other sheets.
            Opens file in write mode if it doesn't exist yet.
            Replaces sheet if same name already exists in the file.
        """
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        self.path = path

        # Use append mode if file exists, otherwise write mode
        file_mode = "a" if os.path.exists(path) else "w"

        # if_sheet_exists is only valid in append mode
        if file_mode == "a":
            self.writer = pd.ExcelWriter(
                path, engine="openpyxl", mode=file_mode, if_sheet_exists="replace"
            )
        else:
            self.writer = pd.ExcelWriter(path, engine="openpyxl", mode=file_mode)

    def write_df(self, sheet: str, df: pd.DataFrame) -> None:
        """
        Write DataFrame to Excel sheet with auto-formatting.

        Args:
            sheet: Name of the Excel sheet to create
            df: pandas DataFrame to write

        Formatting applied:
        - Column widths adjusted to fit header and data content
        - Numeric columns formatted with thousands separators
        - Monetary columns show negatives in parentheses
        - No scientific notation for large numbers
        - Header row styling (bold, background color)
        - Freeze panes at header row
        - Auto-filter enabled on all columns
        """
        # Write DataFrame to sheet
        df.to_excel(self.writer, index=False, sheet_name=sheet)
        ws = self.writer.book[sheet]

        # Define header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths based on BOTH header AND data content
        for i, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)

            # Calculate header width
            header_width = len(str(col))

            # Calculate maximum data width in this column
            max_data_width = 0
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None:
                    # For numbers, estimate formatted length with commas/decimals
                    if isinstance(cell.value, (int, float)):
                        # Estimate formatted string length (e.g., 1234567.89 → "1,234,567.89")
                        if isinstance(cell.value, float):
                            formatted_length = len(f"{cell.value:,.2f}")
                        else:
                            formatted_length = len(f"{cell.value:,}")
                        max_data_width = max(max_data_width, formatted_length)
                    else:
                        # For strings and other types, use actual length
                        max_data_width = max(max_data_width, len(str(cell.value)))

            # Use the larger of header or data width, plus padding
            calculated_width = max(header_width, max_data_width) + self.COLUMN_PADDING
            final_width = max(
                self.MIN_COLUMN_WIDTH, min(self.MAX_COLUMN_WIDTH, calculated_width)
            )
            ws.column_dimensions[col_letter].width = final_width

            # Apply header styling to first row
            header_cell = ws[f"{col_letter}1"]
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = header_alignment

        # Freeze panes at row 2 (freezes header row)
        ws.freeze_panes = "A2"

        # Add auto-filter to all columns
        if ws.max_row > 1:  # Only add filter if there's data
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Format numeric columns to prevent scientific notation
        for i, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)

            # Check if column contains numeric data
            if pd.api.types.is_numeric_dtype(df[col]):
                # Determine if this is a monetary column
                is_monetary = any(
                    keyword in col.lower() for keyword in self.MONETARY_KEYWORDS
                )

                # Select appropriate number format
                number_format = (
                    self.MONETARY_FORMAT if is_monetary else self.NUMERIC_FORMAT
                )

                # Apply format to all data cells in this column (skip header row)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.number_format = number_format

    def close(self) -> None:
        """
        Close the Excel writer and save the file.

        Must be called after all DataFrames have been written.
        """
        self.writer.close()
